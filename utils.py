import openpyxl

def lecture_donnees(file_path='data-folder/data-exemple/exemple.xlsx'):
    """
    Read problem data from an Excel file.

    Parameters:
    - file_path (str): path to the Excel file

    Returns:
    - n (int): total number of time periods
    - tau0 (float): base interest rate per period (as a decimal)
    - placements (list of tuples): each tuple contains (tau_k, d_k, f_k)
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Read n and tau0 from the first two rows
    n = int(sheet.cell(row=1, column=1).value)
    tau0 = float(sheet.cell(row=2, column=1).value) / 100  # Convert percentage to decimal

    # Read investments from rows 3 onward
    placements = []
    row = 3
    while True:
        tauk = sheet.cell(row=row, column=1).value
        dk = sheet.cell(row=row, column=2).value
        fk = sheet.cell(row=row, column=3).value

        if tauk is None or dk is None or fk is None:
            break

        placements.append((float(tauk) / 100, int(dk), int(fk)))
        row += 1

    return n, tau0, placements

def optimiz_coef(t, coef, tau0, placements, chemin):
    """
    Compute the optimal capital coefficient Coef(t) at time t
    using dynamic programming and record the decision path.

    Parameters:
    - t (int): current time step
    - coef (list of float): previously computed Coef values
    - tau0 (float): base interest rate per unit time
    - placements (list of tuples): each placement is (tau_k, d_k, f_k)
    - chemin (list of int): stores the previous time step chosen for each t

    Returns:
    - best (float): the optimal Coef(t) value at time t
    """

    # Option 1: use base interest rate from previous time step (t - 1)
    best = coef[t - 1] * (1 + tau0)

    # Save the origin of the current value: we came from t - 1 via base interest
    chemin[t] = t - 1

    # Option 2: consider specific placements that end exactly at time t
    for tau_k, d_k, f_k in placements:
        if f_k == t:
            # Calculate the value if we use this placement
            val = coef[d_k] * (1 + tau_k)

            # If this value is better than the current best, update
            if val > best:
                best = val
                chemin[t] = d_k  # Record that we came from d_k via this placement

    # Return the optimal capital coefficient at time t
    return best

def reconstruct_path(chemin, end):
    """
    Reconstruct the optimal investment path from the 'chemin' array.

    Parameters:
    - chemin (list): for each t, chemin[t] stores the previous step chosen
    - end (int): the final time step (e.g., n)

    Returns:
    - path (list of tuples): list of (from, to) steps in reverse chronological order
    """
    path = []
    t = end

    while t is not None and chemin[t] is not None:
        prev = chemin[t]
        path.append((prev, t))
        t = prev

    path.reverse()  # so it's from start to end
    return path



def enumerate_paths(n, tau0, placements):
    """
    Enumerates and displays all possible investment paths from time 0 to n,
    along with their corresponding total capital coefficients.

    Parameters:
    - n (int): total number of time periods
    - tau0 (float): base interest rate per period (e.g. 0.009)
    - placements (list of tuples): each tuple is (tau_k, d_k, f_k)
    """
    from collections import defaultdict

    # Step 1: Build the investment graph
    graph = defaultdict(list)

    # Add base transitions: (t, t+1) with base interest rate
    for t in range(n):
        graph[t].append((t + 1, 1 + tau0))

    # Add specific investment arcs (d_k, f_k) with tau_k
    for tau_k, d_k, f_k in placements:
        graph[d_k].append((f_k, 1 + tau_k))

    all_paths = []

    # Step 2: Recursively explore all valid paths from 0 to n
    def explore_paths(t, path, coef):
        if t == n:
            all_paths.append((path[:], coef))
            return

        for next_t, multiplier in graph[t]:
            if next_t <= n:
                path.append((t, next_t))
                explore_paths(next_t, path, coef * multiplier)
                path.pop()

    explore_paths(0, [], 1.0)

    # Step 3: Print all paths and their total coefficients
    for i, (path, coef) in enumerate(all_paths, 1):
        print(f"Path {i}: {path} -> Total Coefficient: {coef:.6f}")



